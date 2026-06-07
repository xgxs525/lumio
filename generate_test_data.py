# -*- coding: utf-8 -*-
"""
测试数据生成脚本
用于生成示例Excel文件以便测试拆分功能
"""

import os
import sys
import random
from datetime import datetime, timedelta

# 导入主脚本中的类
from excel_splitter import ExcelWriter, ColumnSplitRule


def generate_sample_data_by_column(num_rows: int = 1000) -> tuple:
    """生成包含多列的示例数据，适合按列拆分测试"""
    
    departments = ["研发部", "市场部", "销售部", "人力资源部", "财务部", "行政部", "客服部"]
    positions = ["经理", "主管", "专员", "助理", "总监", "高级工程师", "初级工程师"]
    cities = ["北京", "上海", "广州", "深圳", "杭州", "成都", "武汉"]
    
    headers = ["工号", "姓名", "部门", "职位", "城市", "入职日期", "月薪", "绩效评分"]
    data = []
    
    base_date = datetime(2018, 1, 1)
    
    for i in range(num_rows):
        row = {
            "工号": f"EMP{str(i+1).zfill(6)}",
            "姓名": f"员工{i+1}",
            "部门": random.choice(departments),
            "职位": random.choice(positions),
            "城市": random.choice(cities),
            "入职日期": (base_date + timedelta(days=random.randint(0, 2000))).strftime("%Y-%m-%d"),
            "月薪": random.randint(5000, 50000),
            "绩效评分": round(random.uniform(3.0, 5.0), 1)
        }
        data.append(row)
    
    return headers, data


def generate_sample_data_by_row(num_rows: int = 5000) -> tuple:
    """生成包含产品信息的示例数据，适合按行数拆分测试"""
    
    categories = ["电子产品", "服装", "食品", "图书", "家居", "运动", "美妆"]
    statuses = ["在售", "下架", "预售", "缺货"]
    
    headers = ["商品ID", "商品名称", "类别", "品牌", "价格", "库存", "销量", "状态", "添加时间"]
    data = []
    
    base_date = datetime(2023, 1, 1)
    
    for i in range(num_rows):
        row = {
            "商品ID": f"SKU{str(i+1).zfill(8)}",
            "商品名称": f"商品{i+1}",
            "类别": random.choice(categories),
            "品牌": f"品牌{random.randint(1, 20)}",
            "价格": round(random.uniform(9.9, 999.9), 2),
            "库存": random.randint(0, 1000),
            "销量": random.randint(0, 500),
            "状态": random.choice(statuses),
            "添加时间": (base_date + timedelta(days=random.randint(0, 365))).strftime("%Y-%m-%d")
        }
        data.append(row)
    
    return headers, data


def create_test_excel_by_column(output_path: str, num_rows: int = 1000):
    """创建用于按列拆分的测试Excel文件"""
    writer = ExcelWriter(None)
    
    headers, data = generate_sample_data_by_column(num_rows)
    
    # 创建工作簿并写入数据
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "员工信息"
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 写入数据
    for row_idx, row_dict in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header))
    
    # 确保目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    
    wb.save(output_path)
    wb.close()
    
    print(f"已创建测试文件: {output_path}")
    print(f"  - 总行数: {num_rows}")
    print(f"  - 部门分布: 可按'部门'列拆分为7个文件")


def create_test_excel_by_row(output_path: str, num_rows: int = 5000):
    """创建用于按行拆分的测试Excel文件"""
    from openpyxl import Workbook
    
    headers, data = generate_sample_data_by_row(num_rows)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "商品列表"
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 写入数据
    for row_idx, row_dict in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header))
    
    # 确保目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    
    wb.save(output_path)
    wb.close()
    
    print(f"已创建测试文件: {output_path}")
    print(f"  - 总行数: {num_rows}")


def create_test_excel_with_sheets(output_path: str, num_rows_per_sheet: int = 500):
    """创建包含多个工作表的测试Excel文件"""
    from openpyxl import Workbook
    
    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    categories = ["电子产品", "服装", "食品", "图书"]
    
    wb = Workbook()
    
    # 创建第一个工作表（删除默认的）
    if wb.active:
        wb.remove(wb.active)
    
    for month in months:
        ws = wb.create_sheet(title=f"{month}销售")
        
        # 写入表头
        headers = ["日期", "类别", "销售额", "成本", "利润"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 生成该月的数据
        base_date = datetime(2024, int(month.replace("月", "")), 1)
        
        for row_idx in range(num_rows_per_sheet):
            day = random.randint(1, 28)
            sale_date = base_date.replace(day=day) if base_date.day <= 28 else base_date
            
            revenue = random.uniform(1000, 10000)
            cost = revenue * random.uniform(0.5, 0.8)
            
            row = {
                "日期": sale_date.strftime("%Y-%m-%d"),
                "类别": random.choice(categories),
                "销售额": round(revenue, 2),
                "成本": round(cost, 2),
                "利润": round(revenue - cost, 2)
            }
            
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx + 2, column=col_idx, value=row.get(header))
    
    # 确保目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    
    wb.save(output_path)
    wb.close()
    
    print(f"已创建多工作表测试文件: {output_path}")
    print(f"  - 工作表数量: {len(months)}")
    print(f"  - 每个工作表行数: {num_rows_per_sheet}")


def main():
    """主函数"""
    print("=" * 50)
    print("Excel拆分脚本 - 测试数据生成器")
    print("=" * 50)
    
    # 创建测试数据目录
    test_data_dir = "./test_data"
    os.makedirs(test_data_dir, exist_ok=True)
    
    print("\n生成测试数据文件...")
    
    # 1. 用于按列拆分的测试文件
    create_test_excel_by_column(
        os.path.join(test_data_dir, "员工信息_按部门拆分.xlsx"),
        num_rows=500
    )
    
    # 2. 用于按行拆分的测试文件
    create_test_excel_by_row(
        os.path.join(test_data_dir, "商品列表_按行拆分.xlsx"),
        num_rows=2500
    )
    
    # 3. 用于按工作表拆分的测试文件
    create_test_excel_with_sheets(
        os.path.join(test_data_dir, "年度销售报表_多工作表.xlsx"),
        num_rows_per_sheet=200
    )
    
    print("\n" + "=" * 50)
    print("测试数据生成完成！")
    print("=" * 50)
    print("\n测试命令示例:")
    print("-" * 50)
    print("# 按部门列拆分")
    print(f'python excel_splitter.py "{test_data_dir}/员工信息_按部门拆分.xlsx" -c "部门"')
    print()
    print("# 每500行拆分为一个文件")
    print(f'python excel_splitter.py "{test_data_dir}/商品列表_按行拆分.xlsx" --row-count 500')
    print()
    print("# 按工作表拆分")
    print(f'python excel_splitter.py "{test_data_dir}/年度销售报表_多工作表.xlsx" --by-sheet')
    print("-" * 50)


if __name__ == "__main__":
    main()
