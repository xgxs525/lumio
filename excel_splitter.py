# -*- coding: utf-8 -*-
"""
Excel文件拆分脚本
支持按列值、行数、工作表等多种拆分规则
"""

import os
import sys
import logging
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Callable
from dataclasses import dataclass, field
from abc import ABC, abstractmethod
import tempfile
import shutil

# 第三方库
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    xlrd = None  # 可选，用于支持 .xls 格式

try:
    import yaml
except ImportError:
    yaml = None  # 可选，用于支持配置文件


# ==================== 日志配置 ====================

def setup_logging(log_level: str = "INFO", log_file: Optional[str] = None) -> logging.Logger:
    """配置日志系统"""
    logger = logging.getLogger("ExcelSplitter")
    logger.setLevel(getattr(logging, log_level.upper()))
    
    # 清除已有的处理器
    logger.handlers.clear()
    
    # 终端处理器
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%H:%M:%S'
    )
    console_handler.setFormatter(console_format)
    logger.addHandler(console_handler)
    
    # 文件处理器（如果指定）
    if log_file:
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_format = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        file_handler.setFormatter(file_format)
        logger.addHandler(file_handler)
    
    return logger


# ==================== 数据类定义 ====================

@dataclass
class SplitRule:
    """拆分规则基类"""
    name: str
    enabled: bool = True

    @abstractmethod
    def get_split_keys(self, data: List[Dict]) -> List[str]:
        """获取所有需要拆分的键值"""
        pass

    @abstractmethod
    def get_split_key(self, row: Dict) -> str:
        """获取单行数据对应的拆分键"""
        pass


@dataclass
class ColumnSplitRule(SplitRule):
    """按指定列的值拆分"""
    column: str = ""
    
    def get_split_keys(self, data: List[Dict]) -> List[str]:
        keys = set()
        for row in data:
            key = row.get(self.column)
            if key is not None:
                keys.add(str(key))
        return sorted(list(keys))
    
    def get_split_key(self, row: Dict) -> str:
        value = row.get(self.column)
        return str(value) if value is not None else ""


@dataclass
class RowCountSplitRule(SplitRule):
    """按行数拆分"""
    rows_per_file: int = 1000
    
    def get_split_keys(self, data: List[Dict]) -> List[str]:
        total_rows = len(data)
        num_splits = (total_rows + self.rows_per_file - 1) // self.rows_per_file
        return [str(i) for i in range(num_splits)]
    
    def get_split_key(self, row: Dict) -> str:
        return ""  # 需要结合索引使用


@dataclass
class SheetSplitRule(SplitRule):
    """按工作表拆分"""
    
    def get_split_keys(self, data: Any) -> List[str]:
        # data 在这里是工作表名称列表
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], tuple):
                return [str(item[0]) for item in data]
            return [str(item) for item in data]
        return []


@dataclass
class SplitResult:
    """拆分结果"""
    success: bool
    original_file: str
    output_dir: str
    files_created: List[str] = field(default_factory=list)
    error_message: Optional[str] = None
    rows_processed: int = 0
    files_count: int = 0
    duration_seconds: float = 0.0


# ==================== Excel读取器 ====================

class ExcelReader(ABC):
    """Excel读取器抽象基类"""
    
    @abstractmethod
    def read(self, file_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> List[Dict]:
        """读取Excel数据"""
        pass
    
    @abstractmethod
    def get_sheet_names(self, file_path: str) -> List[str]:
        """获取所有工作表名称"""
        pass
    
    @abstractmethod
    def supports_format(self, file_path: str) -> bool:
        """检查是否支持该格式"""
        pass


class OpenpyxlReader(ExcelReader):
    """使用openpyxl读取.xlsx文件"""
    
    def __init__(self, logger: logging.Logger):
        self.logger = logger
    
    def supports_format(self, file_path: str) -> bool:
        return file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xlsm')
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()
    
    def read(self, file_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> List[Dict]:
        """使用迭代器模式读取大型Excel文件以优化内存"""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        try:
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # 获取表头
            headers = []
            header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            for col_idx, cell in enumerate(header_values, start=1):
                headers.append(cell if cell is not None and str(cell).strip() else f"Column_{col_idx}")
            
            self.logger.debug(f"读取到表头: {headers}")
            
            # 迭代读取数据行
            data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                    else:
                        row_dict[f"Column_{col_idx + 1}"] = value
                data.append(row_dict)
                
                # 每10000行输出一次进度
                if row_idx % 10000 == 0:
                    self.logger.info(f"已读取 {row_idx} 行...")
            
            self.logger.info(f"共读取 {len(data)} 行数据")
            return data
            
        finally:
            wb.close()


class XlrdReader(ExcelReader):
    """使用xlrd读取.xls文件"""
    
    def __init__(self, logger: logging.Logger):
        self.logger = logger
        if xlrd is None:
            raise ImportError("需要安装 xlrd 库来读取 .xls 文件: pip install xlrd")
    
    def supports_format(self, file_path: str) -> bool:
        return file_path.lower().endswith('.xls')
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        wb = xlrd.open_workbook(file_path)
        return wb.sheet_names()
    
    def read(self, file_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> List[Dict]:
        wb = xlrd.open_workbook(file_path)
        
        try:
            if sheet_name:
                ws = wb.sheet_by_name(sheet_name)
            else:
                ws = wb.sheet_by_index(0)
            
            # 获取表头
            headers = [ws.cell_value(header_row - 1, col) for col in range(ws.ncols)]
            headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(headers)]
            
            self.logger.debug(f"读取到表头: {headers}")
            
            # 读取数据行
            data = []
            for row_idx in range(header_row, ws.nrows):
                row_dict = {}
                for col_idx in range(ws.ncols):
                    row_dict[headers[col_idx]] = ws.cell_value(row_idx, col_idx)
                data.append(row_dict)
                
                if row_idx % 10000 == 0:
                    self.logger.info(f"已读取 {row_idx} 行...")
            
            self.logger.info(f"共读取 {len(data)} 行数据")
            return data
            
        finally:
            wb.release_resources()


# ==================== Excel写入器 ====================

class ExcelWriter:
    """Excel写入器"""
    
    def __init__(self, logger: logging.Logger):
        self.logger = logger
    
    def write(self, data: List[Dict], output_path: str, sheet_name: str = "Sheet1") -> int:
        """写入数据到Excel文件"""
        if not data:
            self.logger.warning(f"没有数据写入到 {output_path}")
            return 0
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel工作表名称最多31个字符
        
        # 写入表头
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 写入数据行
        for row_idx, row_dict in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header))
            
            if row_idx % 10000 == 0:
                self.logger.debug(f"已写入 {row_idx} 行...")
        
        wb.save(output_path)
        wb.close()
        
        self.logger.info(f"已保存 {output_path}，共 {len(data)} 行")
        return len(data)


# ==================== 拆分策略 ====================

class SplitStrategy(ABC):
    """拆分策略抽象基类"""
    
    @abstractmethod
    def split(
        self,
        data: List[Dict],
        rule: SplitRule,
        output_dir: str,
        original_filename: str,
        writer: ExcelWriter,
        logger: logging.Logger
    ) -> List[str]:
        """执行拆分，返回生成的文件路径列表"""
        pass


class ColumnSplitStrategy(SplitStrategy):
    """按列值拆分策略"""
    
    def split(
        self,
        data: List[Dict],
        rule: SplitRule,
        output_dir: str,
        original_filename: str,
        writer: ExcelWriter,
        logger: logging.Logger
    ) -> List[str]:
        if not isinstance(rule, ColumnSplitRule):
            raise ValueError("需要 ColumnSplitRule")
        
        column = rule.column
        if not column:
            raise ValueError("未指定拆分列")
        
        # 检查列是否存在
        if data and column not in data[0]:
            raise ValueError(f"指定的拆分列 '{column}' 不存在于数据中")
        
        # 按列值分组
        groups: Dict[str, List[Dict]] = {}
        for row in data:
            key = rule.get_split_key(row)
            if key not in groups:
                groups[key] = []
            groups[key].append(row)
        
        # 创建输出文件
        base_name = Path(original_filename).stem
        output_files = []
        
        for key, rows in sorted(groups.items()):
            safe_key = self._sanitize_filename(str(key))
            output_path = os.path.join(output_dir, f"{base_name}_{safe_key}.xlsx")
            writer.write(rows, output_path)
            output_files.append(output_path)
            logger.info(f"  列值 '{key}' -> {len(rows)} 行 -> {output_path}")
        
        return output_files
    
    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        """清理文件名中的非法字符"""
        illegal_chars = '<>:"/\\|?*'
        for char in illegal_chars:
            filename = filename.replace(char, '_')
        return filename[:100]  # 限制长度


class RowCountSplitStrategy(SplitStrategy):
    """按行数拆分策略"""
    
    def split(
        self,
        data: List[Dict],
        rule: SplitRule,
        output_dir: str,
        original_filename: str,
        writer: ExcelWriter,
        logger: logging.Logger
    ) -> List[str]:
        if not isinstance(rule, RowCountSplitRule):
            raise ValueError("需要 RowCountSplitRule")
        
        rows_per_file = rule.rows_per_file
        if rows_per_file <= 0:
            raise ValueError("每文件行数必须大于0")
        
        base_name = Path(original_filename).stem
        output_files = []
        total_rows = len(data)
        
        for start_idx in range(0, total_rows, rows_per_file):
            end_idx = min(start_idx + rows_per_file, total_rows)
            chunk_data = data[start_idx:end_idx]
            
            part_num = start_idx // rows_per_file + 1
            output_path = os.path.join(output_dir, f"{base_name}_part{part_num:03d}.xlsx")
            writer.write(chunk_data, output_path)
            output_files.append(output_path)
            logger.info(f"  部分 {part_num} ({start_idx+1}-{end_idx}) -> {len(chunk_data)} 行")
        
        return output_files


class SheetSplitStrategy(SplitStrategy):
    """按工作表拆分策略"""
    
    def split(
        self,
        data: Any,
        rule: SplitRule,
        output_dir: str,
        original_filename: str,
        writer: ExcelWriter,
        logger: logging.Logger
    ) -> List[str]:
        # data 在这里是 (sheet_name, data) 的元组列表
        if not isinstance(data, list):
            raise ValueError("SheetSplitStrategy 需要工作表数据列表")
        
        base_name = Path(original_filename).stem
        output_files = []
        
        for sheet_name, sheet_data in data:
            safe_name = ColumnSplitStrategy._sanitize_filename(str(sheet_name))
            output_path = os.path.join(output_dir, f"{base_name}_{safe_name}.xlsx")
            writer.write(sheet_data, output_path)
            output_files.append(output_path)
            logger.info(f"  工作表 '{sheet_name}' -> {len(sheet_data)} 行")
        
        return output_files


# ==================== Excel拆分器主类 ====================

class ExcelSplitter:
    """Excel文件拆分器"""
    
    STRATEGIES = {
        "column": ColumnSplitStrategy(),
        "row_count": RowCountSplitStrategy(),
        "sheet": SheetSplitStrategy()
    }
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger or setup_logging()
        self.reader: Optional[ExcelReader] = None
        self.writer = ExcelWriter(self.logger)
    
    def _get_reader(self, file_path: str) -> ExcelReader:
        """根据文件格式选择合适的读取器"""
        if OpenpyxlReader(None).supports_format(file_path):
            return OpenpyxlReader(self.logger)
        elif xlrd is not None and XlrdReader(None).supports_format(file_path):
            return XlrdReader(self.logger)
        else:
            raise ValueError(f"不支持的文件格式: {file_path}")
    
    def _validate_file(self, file_path: str) -> None:
        """验证文件是否存在且可读"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if not os.path.isfile(file_path):
            raise ValueError(f"不是有效的文件: {file_path}")
        
        # 检查文件扩展名
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls', '.xlsm']:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx, .xls, .xlsm")
        
        self.logger.debug(f"文件验证通过: {file_path}")
    
    def split_by_column(
        self,
        input_file: str,
        output_dir: str,
        column: str,
        header_row: int = 1,
        create_subdir: bool = True
    ) -> SplitResult:
        """按指定列的值拆分文件"""
        return self._split(
            input_file=input_file,
            output_dir=output_dir,
            rule=ColumnSplitRule(name="column_split", column=column),
            header_row=header_row,
            create_subdir=create_subdir
        )
    
    def split_by_row_count(
        self,
        input_file: str,
        output_dir: str,
        rows_per_file: int,
        header_row: int = 1,
        create_subdir: bool = True
    ) -> SplitResult:
        """按指定行数拆分文件"""
        return self._split(
            input_file=input_file,
            output_dir=output_dir,
            rule=RowCountSplitRule(name="row_count_split", rows_per_file=rows_per_file),
            header_row=header_row,
            create_subdir=create_subdir
        )
    
    def split_by_sheet(
        self,
        input_file: str,
        output_dir: str,
        header_row: int = 1,
        create_subdir: bool = True
    ) -> SplitResult:
        """按工作表拆分文件"""
        return self._split(
            input_file=input_file,
            output_dir=output_dir,
            rule=SheetSplitRule(name="sheet_split"),
            header_row=header_row,
            create_subdir=create_subdir
        )
    
    def _split(
        self,
        input_file: str,
        output_dir: str,
        rule: SplitRule,
        header_row: int = 1,
        create_subdir: bool = True
    ) -> SplitResult:
        """通用拆分方法"""
        start_time = datetime.now()
        result = SplitResult(
            success=False,
            original_file=input_file,
            output_dir=output_dir
        )
        
        try:
            # 验证文件
            self._validate_file(input_file)
            
            # 获取读取器
            self.reader = self._get_reader(input_file)
            
            # 确定策略类型
            strategy_type = rule.name.split('_')[0]  # 从规则名称获取策略类型
            
            # 创建输出目录
            if create_subdir:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                subdir_name = f"{Path(input_file).stem}_split_{timestamp}"
                output_dir = os.path.join(output_dir, subdir_name)
            
            os.makedirs(output_dir, exist_ok=True)
            result.output_dir = output_dir
            self.logger.info(f"输出目录: {output_dir}")
            
            # 根据策略类型执行不同的读取和拆分逻辑
            if strategy_type == "column" or strategy_type == "row":
                # 读取数据
                data = self.reader.read(input_file, header_row=header_row)
                result.rows_processed = len(data)
                
                # 执行拆分
                strategy = self.STRATEGIES.get(
                    "column" if strategy_type == "column" else "row_count"
                )
                result.files_created = strategy.split(
                    data=data,
                    rule=rule,
                    output_dir=output_dir,
                    original_filename=input_file,
                    writer=self.writer,
                    logger=self.logger
                )
                
            elif strategy_type == "sheet":
                # 按工作表读取和拆分
                sheet_names = self.reader.get_sheet_names(input_file)
                self.logger.info(f"发现 {len(sheet_names)} 个工作表: {sheet_names}")
                
                sheets_data = []
                for sheet_name in sheet_names:
                    data = self.reader.read(input_file, sheet_name=sheet_name, header_row=header_row)
                    sheets_data.append((sheet_name, data))
                
                total_rows = sum(len(d) for _, d in sheets_data)
                result.rows_processed = total_rows
                
                strategy = self.STRATEGIES.get("sheet")
                result.files_created = strategy.split(
                    data=sheets_data,
                    rule=rule,
                    output_dir=output_dir,
                    original_filename=input_file,
                    writer=self.writer,
                    logger=self.logger
                )
            
            result.files_count = len(result.files_created)
            result.success = True
            
        except FileNotFoundError as e:
            result.error_message = f"文件错误: {str(e)}"
            self.logger.error(result.error_message)
            
        except ValueError as e:
            result.error_message = f"数据错误: {str(e)}"
            self.logger.error(result.error_message)
            
        except Exception as e:
            result.error_message = f"未知错误: {str(e)}"
            self.logger.exception("拆分过程发生异常")
            
        finally:
            result.duration_seconds = (datetime.now() - start_time).total_seconds()
        
        return result
    
    def generate_report(self, result: SplitResult) -> str:
        """生成结果报告"""
        report_lines = [
            "=" * 50,
            "Excel拆分结果报告",
            "=" * 50,
            f"原始文件: {result.original_file}",
            f"输出目录: {result.output_dir}",
            f"执行状态: {'成功' if result.success else '失败'}",
            f"处理行数: {result.rows_processed}",
            f"生成文件数: {result.files_count}",
            f"耗时: {result.duration_seconds:.2f}秒",
        ]
        
        if result.files_created:
            report_lines.append("\n生成的文件:")
            for i, file_path in enumerate(result.files_created, 1):
                file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                report_lines.append(f"  {i}. {file_path} ({file_size:,} bytes)")
        
        if result.error_message:
            report_lines.append(f"\n错误信息: {result.error_message}")
        
        report_lines.append("=" * 50)
        
        return "\n".join(report_lines)


# ==================== 命令行接口 ====================

def parse_args():
    """解析命令行参数"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Excel文件拆分工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  # 按列值拆分
  python excel_splitter.py input.xlsx -c "部门" -o output/

  # 按行数拆分
  python excel_splitter.py input.xlsx --row-count 5000 -o output/

  # 按工作表拆分
  python excel_splitter.py input.xlsx --by-sheet -o output/

  # 使用配置文件
  python excel_splitter.py --config config.yaml
        """
    )
    
    parser.add_argument('input_file', nargs='?', help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', default='./output', help='输出目录 (默认: ./output)')
    parser.add_argument('-c', '--column', help='按指定列的值拆分')
    parser.add_argument('--row-count', type=int, help='每个文件的行数')
    parser.add_argument('--by-sheet', action='store_true', help='按工作表拆分')
    parser.add_argument('--config', help='配置文件路径 (YAML格式)')
    parser.add_argument('--log-level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                        help='日志级别 (默认: INFO)')
    parser.add_argument('--log-file', help='日志文件路径')
    parser.add_argument('--no-subdir', action='store_true', help='不在输出目录中创建时间戳子目录')
    
    return parser.parse_args()


def load_config(config_path: str, logger: logging.Logger) -> Dict[str, Any]:
    """加载配置文件"""
    if yaml is None:
        logger.error("需要安装 pyyaml 库来使用配置文件: pip install pyyaml")
        sys.exit(1)
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        logger.info(f"已加载配置文件: {config_path}")
        return config
    except Exception as e:
        logger.error(f"加载配置文件失败: {e}")
        sys.exit(1)


def main():
    """主函数"""
    args = parse_args()
    
    # 设置日志
    logger = setup_logging(args.log_level, args.log_file)
    
    # 创建拆分器实例
    splitter = ExcelSplitter(logger)
    
    # 如果指定了配置文件
    if args.config:
        config = load_config(args.config, logger)
        
        # 处理配置文件中的多个任务
        tasks = config.get('tasks', [config])
        
        for task in tasks:
            input_file = task.get('input_file')
            if not input_file:
                logger.error("配置文件中未指定 input_file")
                continue
            
            output_dir = task.get('output_dir', args.output)
            rule_type = task.get('rule_type')
            create_subdir = not args.no_subdir
            
            logger.info(f"\n开始处理: {input_file}")
            
            if rule_type == 'column':
                result = splitter.split_by_column(
                    input_file, output_dir, task['column'], create_subdir=create_subdir
                )
            elif rule_type == 'row_count':
                result = splitter.split_by_row_count(
                    input_file, output_dir, task['rows_per_file'], create_subdir=create_subdir
                )
            elif rule_type == 'sheet':
                result = splitter.split_by_sheet(
                    input_file, output_dir, create_subdir=create_subdir
                )
            else:
                logger.error(f"未知的规则类型: {rule_type}")
                continue
            
            logger.info("\n" + splitter.generate_report(result))
        
        return
    
    # 命令行参数模式
    if not args.input_file:
        logger.error("请指定输入文件或配置文件")
        sys.exit(1)
    
    # 确定拆分方式
    create_subdir = not args.no_subdir
    
    if args.by_sheet:
        result = splitter.split_by_sheet(args.input_file, args.output, create_subdir=create_subdir)
    elif args.row_count:
        result = splitter.split_by_row_count(
            args.input_file, args.output, args.row_count, create_subdir=create_subdir
        )
    elif args.column:
        result = splitter.split_by_column(
            args.input_file, args.output, args.column, create_subdir=create_subdir
        )
    else:
        logger.error("请指定拆分方式: -c (列名), --row-count (行数), 或 --by-sheet")
        sys.exit(1)
    
    # 输出报告
    report = splitter.generate_report(result)
    print("\n" + report)
    
    # 返回退出码
    sys.exit(0 if result.success else 1)


if __name__ == "__main__":
    main()
