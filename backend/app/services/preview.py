import openpyxl
import xlrd


def preview_split(
    input_path: str,
    split_type: str,
    *,
    column: str | None = None,
    rows_per_file: int | None = None,
    header_row: int = 1,
) -> dict:
    preview_result: dict = {'success': True, 'splitType': split_type}

    if split_type == 'column':
        if input_path.endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
            col_idx = next((i + 1 for i, h in enumerate(headers) if str(h) == column), None)
            if col_idx is None:
                wb.close()
                raise ValueError(f'未找到列: {column}')
            value_counts: dict[str, int] = {}
            total_rows = 0
            for row in ws.iter_rows(min_row=header_row + 1):
                cell_value = row[col_idx - 1].value
                if cell_value:
                    key = str(cell_value)
                    value_counts[key] = value_counts.get(key, 0) + 1
                    total_rows += 1
            wb.close()
        else:
            wb = xlrd.open_workbook(input_path)
            ws = wb.sheet_by_index(0)
            headers = [ws.cell_value(header_row - 1, col) for col in range(ws.ncols)]
            col_idx = next((i for i, h in enumerate(headers) if str(h) == column), None)
            if col_idx is None:
                raise ValueError(f'未找到列: {column}')
            value_counts = {}
            total_rows = 0
            for row_idx in range(header_row, ws.nrows):
                cell_value = ws.cell_value(row_idx, col_idx)
                if cell_value:
                    key = str(cell_value)
                    value_counts[key] = value_counts.get(key, 0) + 1
                    total_rows += 1

        preview_result.update({
            'column': column,
            'valueCounts': value_counts,
            'totalRows': total_rows,
            'fileCount': len(value_counts),
            'files': [
                {'name': f'{key}.xlsx', 'rows': count}
                for key, count in sorted(value_counts.items(), key=lambda x: -x[1])
            ],
        })

    elif split_type == 'row_count':
        if input_path.endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(input_path, read_only=True)
            ws = wb.active
            total_rows = sum(1 for _ in ws.iter_rows(min_row=header_row + 1))
            wb.close()
        else:
            wb = xlrd.open_workbook(input_path)
            ws = wb.sheet_by_index(0)
            total_rows = max(0, ws.nrows - header_row)

        file_count = (total_rows + rows_per_file - 1) // rows_per_file
        preview_result.update({
            'rowsPerFile': rows_per_file,
            'totalRows': total_rows,
            'fileCount': file_count,
            'files': [
                {'name': f'part_{i + 1}.xlsx', 'rows': min(rows_per_file, total_rows - i * rows_per_file)}
                for i in range(file_count)
            ],
        })

    elif split_type == 'sheet':
        if input_path.endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(input_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:
            wb = xlrd.open_workbook(input_path)
            sheet_names = wb.sheet_names()

        preview_result.update({
            'fileCount': len(sheet_names),
            'files': [{'name': f'{name}.xlsx', 'rows': '全部'} for name in sheet_names],
        })

    return preview_result
