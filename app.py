# -*- coding: utf-8 -*-
"""
序办 - Flask Web服务
"""

import os
import sys
from datetime import datetime
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template, redirect
from werkzeug.security import safe_join
from werkzeug.utils import secure_filename
import threading
import uuid

# 导入Excel拆分器
from excel_splitter import ExcelSplitter, setup_logging

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 限制500MB
app.config['UPLOAD_FOLDER'] = './uploads'
app.config['OUTPUT_FOLDER'] = './outputs'
app.config['TEMPLATE_FOLDER'] = './uploads/templates'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # 禁用缓存
app.jinja_env.cache = {}  # 禁用Jinja2缓存

# 确保目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
os.makedirs(app.config['TEMPLATE_FOLDER'], exist_ok=True)

# 日志配置
logger = setup_logging('INFO')
splitter = ExcelSplitter(logger)

# 任务状态存储
tasks = {}
ALLOWED_SPLIT_TYPES = {'column', 'row_count', 'sheet'}
ALLOWED_TEMPLATE_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.csv', '.doc', '.docx', '.ppt', '.pptx', '.pdf', '.txt', '.md'}
ALLOWED_UPLOAD_EXTENSIONS = {
    '.xlsx', '.xls', '.xlsm', '.csv',
    '.doc', '.docx', '.ppt', '.pptx', '.pdf',
    '.txt', '.md', '.png', '.jpg', '.jpeg'
}


def get_json_payload():
    """获取JSON请求体，避免空请求体导致AttributeError。"""
    return request.get_json(silent=True) or {}


def coerce_positive_int(value, field_name: str) -> int:
    """把前端传来的数字/字符串统一转成正整数。"""
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f'{field_name}必须是大于0的整数')

    if number <= 0:
        raise ValueError(f'{field_name}必须是大于0的整数')

    return number


def is_uploaded_file(file_path: str) -> bool:
    """限制Web接口只能处理上传目录中的文件。"""
    try:
        upload_dir = Path(app.config['UPLOAD_FOLDER']).resolve()
        target_path = Path(file_path).resolve()
        target_path.relative_to(upload_dir)
    except (TypeError, ValueError, OSError):
        return False

    return target_path.is_file()


def run_split_task(task_id: str, input_path: str, output_dir: str, split_type: str, **kwargs):
    """在后台线程中执行拆分任务"""
    try:
        tasks[task_id]['status'] = 'running'
        tasks[task_id]['message'] = '开始拆分...'
        
        if split_type == 'column':
            result = splitter.split_by_column(
                input_path, output_dir, kwargs['column'],
                header_row=kwargs.get('header_row', 1),
                create_subdir=False
            )
        elif split_type == 'row_count':
            result = splitter.split_by_row_count(
                input_path, output_dir, int(kwargs['rows_per_file']),
                header_row=kwargs.get('header_row', 1),
                create_subdir=False
            )
        elif split_type == 'sheet':
            result = splitter.split_by_sheet(
                input_path, output_dir,
                header_row=kwargs.get('header_row', 1),
                create_subdir=False
            )
        else:
            raise ValueError(f"未知拆分类型: {split_type}")
        
        tasks[task_id]['status'] = 'completed'
        tasks[task_id]['success'] = result.success
        tasks[task_id]['message'] = splitter.generate_report(result)
        tasks[task_id]['output_dir'] = result.output_dir
        tasks[task_id]['files'] = result.files_created
        tasks[task_id]['files_count'] = result.files_count
        tasks[task_id]['rows_processed'] = result.rows_processed
        
        if not result.success:
            tasks[task_id]['error'] = result.error_message
            
    except Exception as e:
        tasks[task_id]['status'] = 'failed'
        tasks[task_id]['error'] = str(e)
        logger.exception(f"任务 {task_id} 执行失败")


@app.route('/')
def index():
    """首页（欢迎页）"""
    return render_template('home.html')

@app.route('/tutorial')
def tutorial():
    """教程页面"""
    return render_template('tutorial.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传办公文件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    
    # 检查文件扩展名（使用原始文件名检查）
    original_filename = file.filename
    ext = os.path.splitext(original_filename)[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return jsonify({'success': False, 'error': f'不支持的文件格式: {ext}'}), 400
    
    # 使用UUID生成唯一文件名，避免中文文件名问题
    unique_filename = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
    file.save(file_path)
    
    logger.info(f"文件上传成功: {file_path}")
    
    return jsonify({
        'success': True,
        'filename': original_filename,  # 返回原始文件名给前端显示
        'filepath': file_path,
        'size': os.path.getsize(file_path)
    })


@app.route('/api/preview', methods=['POST'])
def preview_split():
    """预览拆分结果（不实际拆分，只计算会生成多少个文件）"""
    data = get_json_payload()
    
    input_path = data.get('filepath')
    split_type = data.get('splitType', data.get('split_type'))  # column, row_count, sheet
    column = data.get('column')
    rows_per_file = data.get('rowsPerFile', data.get('rows_per_file'))
    try:
        header_row = coerce_positive_int(data.get('headerRow', data.get('header_row', 1)), '表头行')
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    
    if not input_path or not os.path.exists(input_path):
        return jsonify({'success': False, 'error': '文件不存在'}), 400

    if not is_uploaded_file(input_path):
        return jsonify({'success': False, 'error': '只能处理已上传的文件'}), 400

    if split_type not in ALLOWED_SPLIT_TYPES:
        return jsonify({'success': False, 'error': '未知拆分类型'}), 400
    
    if split_type == 'column' and not column:
        return jsonify({'success': False, 'error': '请指定拆分列'}), 400
    
    if split_type == 'row_count':
        try:
            rows_per_file = coerce_positive_int(rows_per_file, '每文件行数')
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400
    
    try:
        preview_result = {'success': True, 'splitType': split_type}
        
        if split_type == 'column':
            # 按列值预览：统计每个值有多少行
            wb = None
            if input_path.endswith('.xlsx') or input_path.endswith('.xlsm'):
                import openpyxl
                wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
                ws = wb.active
                
                # 找到列索引
                headers = [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
                col_idx = None
                for i, h in enumerate(headers):
                    if str(h) == column:
                        col_idx = i + 1  # openpyxl使用1-based索引
                        break
                
                if col_idx is None:
                    return jsonify({'success': False, 'error': f'未找到列: {column}'}), 400
                
                # 统计每个值的行数
                value_counts = {}
                total_rows = 0
                for row in ws.iter_rows(min_row=header_row + 1):
                    cell_value = row[col_idx - 1].value
                    if cell_value:
                        key = str(cell_value)
                        value_counts[key] = value_counts.get(key, 0) + 1
                        total_rows += 1
                
                preview_result['column'] = column
                preview_result['valueCounts'] = value_counts
                preview_result['totalRows'] = total_rows
                preview_result['fileCount'] = len(value_counts)
                preview_result['files'] = [
                    {'name': f'{key}.xlsx', 'rows': count}
                    for key, count in sorted(value_counts.items(), key=lambda x: -x[1])
                ]
                
            else:
                import xlrd
                wb = xlrd.open_workbook(input_path)
                ws = wb.sheet_by_index(0)
                
                # 找到列索引
                headers = [ws.cell_value(header_row - 1, col) for col in range(ws.ncols)]
                col_idx = None
                for i, h in enumerate(headers):
                    if str(h) == column:
                        col_idx = i
                        break
                
                if col_idx is None:
                    return jsonify({'success': False, 'error': f'未找到列: {column}'}), 400
                
                # 统计每个值的行数
                value_counts = {}
                total_rows = 0
                for row_idx in range(header_row, ws.nrows):
                    cell_value = ws.cell_value(row_idx, col_idx)
                    if cell_value:
                        key = str(cell_value)
                        value_counts[key] = value_counts.get(key, 0) + 1
                        total_rows += 1
                
                preview_result['column'] = column
                preview_result['valueCounts'] = value_counts
                preview_result['totalRows'] = total_rows
                preview_result['fileCount'] = len(value_counts)
                preview_result['files'] = [
                    {'name': f'{key}.xlsx', 'rows': count}
                    for key, count in sorted(value_counts.items(), key=lambda x: -x[1])
                ]
            
            if hasattr(wb, 'close'):
                wb.close()
                
        elif split_type == 'row_count':
            # 按行数预览：计算会生成多少个文件
            wb = None
            if input_path.endswith('.xlsx') or input_path.endswith('.xlsm'):
                import openpyxl
                wb = openpyxl.load_workbook(input_path, read_only=True)
                ws = wb.active
                
                # 计算总行数（不包括表头）
                total_rows = 0
                for _ in ws.iter_rows(min_row=header_row + 1):
                    total_rows += 1
                
                file_count = (total_rows + rows_per_file - 1) // rows_per_file  # 向上取整
                
                preview_result['rowsPerFile'] = rows_per_file
                preview_result['totalRows'] = total_rows
                preview_result['fileCount'] = file_count
                preview_result['files'] = [
                    {'name': f'part_{i+1}.xlsx', 'rows': min(rows_per_file, total_rows - i * rows_per_file)}
                    for i in range(file_count)
                ]
                
            else:
                import xlrd
                wb = xlrd.open_workbook(input_path)
                ws = wb.sheet_by_index(0)
                
                total_rows = max(0, ws.nrows - header_row)  # 减去表头行
                file_count = (total_rows + rows_per_file - 1) // rows_per_file
                
                preview_result['rowsPerFile'] = rows_per_file
                preview_result['totalRows'] = total_rows
                preview_result['fileCount'] = file_count
                preview_result['files'] = [
                    {'name': f'part_{i+1}.xlsx', 'rows': min(rows_per_file, total_rows - i * rows_per_file)}
                    for i in range(file_count)
                ]
            
            if hasattr(wb, 'close'):
                wb.close()
                
        elif split_type == 'sheet':
            # 按工作表预览：列出所有工作表
            wb = None
            if input_path.endswith('.xlsx') or input_path.endswith('.xlsm'):
                import openpyxl
                wb = openpyxl.load_workbook(input_path, read_only=True)
                sheet_names = wb.sheetnames
                
                preview_result['fileCount'] = len(sheet_names)
                preview_result['files'] = [
                    {'name': f'{name}.xlsx', 'rows': '全部'}
                    for name in sheet_names
                ]
                
            else:
                import xlrd
                wb = xlrd.open_workbook(input_path)
                sheet_names = wb.sheet_names()
                
                preview_result['fileCount'] = len(sheet_names)
                preview_result['files'] = [
                    {'name': f'{name}.xlsx', 'rows': '全部'}
                    for name in sheet_names
                ]
            
            if hasattr(wb, 'close'):
                wb.close()
        
        return jsonify(preview_result)
        
    except Exception as e:
        logger.exception(f"预览失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/split', methods=['POST'])
def split_file():
    """执行拆分任务"""
    data = get_json_payload()
    
    input_path = data.get('filepath')
    split_type = data.get('splitType', data.get('split_type'))  # column, row_count, sheet
    column = data.get('column')
    rows_per_file = data.get('rowsPerFile', data.get('rows_per_file'))
    try:
        header_row = coerce_positive_int(data.get('headerRow', data.get('header_row', 1)), '表头行')
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    
    if not input_path or not os.path.exists(input_path):
        return jsonify({'success': False, 'error': '文件不存在'}), 400

    if not is_uploaded_file(input_path):
        return jsonify({'success': False, 'error': '只能处理已上传的文件'}), 400

    if split_type not in ALLOWED_SPLIT_TYPES:
        return jsonify({'success': False, 'error': '未知拆分类型'}), 400
    
    if split_type == 'column' and not column:
        return jsonify({'success': False, 'error': '请指定拆分列'}), 400
    
    if split_type == 'row_count':
        try:
            rows_per_file = coerce_positive_int(rows_per_file, '每文件行数')
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400
    
    # 创建任务
    task_id = uuid.uuid4().hex
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(
        app.config['OUTPUT_FOLDER'],
        f"{Path(input_path).stem}_split_{timestamp}"
    )
    os.makedirs(output_dir, exist_ok=True)
    
    tasks[task_id] = {
        'status': 'pending',
        'message': '等待处理...',
        'input_file': input_path,
        'output_dir': output_dir,
        'split_type': split_type,
        'files': [],
        'success': False
    }
    
    # 启动后台任务
    thread = threading.Thread(
        target=run_split_task,
        args=(task_id, input_path, output_dir, split_type),
        kwargs={'column': column, 'rows_per_file': rows_per_file, 'header_row': header_row}
    )
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'taskId': task_id,
        'outputDir': output_dir
    })


@app.route('/api/task/<task_id>')
def get_task_status(task_id):
    """获取任务状态"""
    if task_id not in tasks:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    return jsonify({
        'success': True,
        'task': tasks[task_id]
    })


@app.route('/api/download/<task_id>/<filename>')
def download_file(task_id, filename):
    """下载生成的文件"""
    if task_id not in tasks:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    task = tasks[task_id]
    file_path = safe_join(task['output_dir'], filename)
    
    if file_path is None or not os.path.isfile(file_path):
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    return send_file(file_path, as_attachment=True)


@app.route('/api/download-all/<task_id>')
def download_all_files(task_id):
    """打包下载所有生成的文件"""
    if task_id not in tasks:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    import zipfile
    from io import BytesIO
    
    task = tasks[task_id]
    output_dir = task['output_dir']
    
    if not task.get('files'):
        return jsonify({'success': False, 'error': '没有生成的文件'}), 400
    
    # 创建ZIP文件
    memory_file = BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file_path in task['files']:
            if os.path.exists(file_path):
                zf.write(file_path, os.path.basename(file_path))
    
    memory_file.seek(0)
    
    zip_name = f"{Path(task['input_file']).stem}_split_results.zip"
    
    return send_file(
        memory_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_name
    )


# 工具中心路由
@app.route('/tools')
def tools():
    return render_template('tools.html')

@app.route('/split')
def split():
    return render_template('index.html')

@app.route('/tools/split')
def tool_split():
    return render_template('split.html')

@app.route('/dedup')
def dedup():
    return render_template('dedup.html')

@app.route('/convert')
def convert():
    return render_template('convert.html')

@app.route('/beautify')
def beautify():
    return render_template('beautify.html')

@app.route('/formulas')
def formulas():
    return render_template('formulas.html')

@app.route('/statistics')
def statistics():
    return render_template('statistics.html')

@app.route('/charts')
def charts():
    return render_template('charts.html')

@app.route('/batch')
def batch():
    return render_template('batch.html')

@app.route('/clean')
def clean():
    return render_template('clean.html')

@app.route('/tools/clean')
def tool_clean():
    return render_template('clean.html')

@app.route('/tools/dedup')
def tool_dedup():
    return render_template('dedup.html')

@app.route('/tools/convert')
def tool_convert():
    return render_template('convert.html')

@app.route('/tools/beautify')
def tool_beautify():
    return render_template('beautify.html')

@app.route('/tools/formulas')
def tool_formulas():
    return render_template('formulas.html')

@app.route('/tools/statistics')
def tool_statistics():
    return render_template('statistics.html')

@app.route('/tools/charts')
def tool_charts():
    return render_template('charts.html')

@app.route('/tools/batch')
def tool_batch():
    return render_template('batch.html')

# AI 办公助手路由
@app.route('/ai')
def ai_excel():
    return render_template('ai.html')

# 模板中心路由
@app.route('/templates')
def templates():
    return render_template('templates.html')

# 解决方案路由
@app.route('/solutions')
def solutions():
    return render_template('solutions.html')

# 资源中心路由
@app.route('/resources')
def resources():
    return render_template('tutorial.html')

def render_resource_page(title, subtitle, sections):
    return render_template(
        'legal.html',
        title=title,
        subtitle=subtitle,
        sections=sections
    )

@app.route('/resources/cases')
def resource_cases():
    return render_resource_page(
        '案例中心',
        '展示电商、财务、销售、人事、仓储和团队协作中的真实办公场景，帮助用户理解序办如何落地到业务流程。',
        [
            ('电商运营案例', '按地区拆分订单、清洗商品数据、汇总销售日报，并沉淀为团队可复用模板。'),
            ('财务办公案例', '整理账单、费用明细、发票资料和对账差异，减少重复核对工作。'),
            ('销售管理案例', '管理客户表、跟进记录、回款计划和区域业绩，用数据看板复盘销售过程。'),
            ('团队协作案例', '把文档、表格、知识库、云盘和自动化任务连接起来，形成统一工作台。')
        ]
    )

@app.route('/resources/help')
def resource_help():
    return render_resource_page(
        '帮助文档',
        '集中说明账号、文件上传、数据处理、模板、云盘、会员和支持相关问题。',
        [
            ('账号与登录', '支持账号密码、手机号验证码和第三方方式登录，登录后可管理账号安全与团队资料。'),
            ('文件上传', '上传办公文件后先停留在当前页面，确认需求或点击开始处理后再执行任务。'),
            ('模板与云盘', '模板可收藏、上传、下载和复用；云盘用于保存文件、处理结果和知识库附件。'),
            ('会员与支持', '会员权益、额度、订单和支持入口都可以在工作台或账号菜单中查看。')
        ]
    )

@app.route('/resources/updates')
def resource_updates():
    return render_resource_page(
        '更新日志',
        '记录序办产品能力、页面体验和平台服务的近期更新。',
        [
            ('平台导航升级', '顶部菜单升级为首页、产品、解决方案、模板中心、资源中心、价格、工作台。'),
            ('云盘与知识库', '新增云盘文件管理、知识库新建、知识库设置和团队资料沉淀入口。'),
            ('账号体验优化', '登录后显示头像菜单，支持语言、外观、切换账号、账号与安全和注销账号。'),
            ('资源页面拆分', '案例中心、帮助文档、更新日志、API 文档、安全合规和白皮书改为独立页面。')
        ]
    )

@app.route('/resources/api')
def resource_api():
    return render_resource_page(
        'API 文档',
        '面向企业客户和开发者的集成说明，用于后续接入文件处理、数据分析、自动化流程和账号体系。',
        [
            ('接口概览', '提供文件上传、任务创建、任务状态查询、结果下载和模板管理等能力。'),
            ('认证方式', '企业接入可使用项目密钥、团队权限和操作审计，保障接口调用安全。'),
            ('任务回调', '处理完成后可通过回调通知业务系统，方便接入审批、消息和数据流转。'),
            ('开发支持', '企业版可提供专属接口联调、字段映射和自动化流程定制说明。')
        ]
    )

@app.route('/resources/security')
def resource_security():
    return render_resource_page(
        '安全与合规',
        '说明序办在数据安全、权限管理、文件隔离和企业合规方面的设计。',
        [
            ('数据隔离', '上传文件、处理结果、模板和知识库资料按账号和团队空间隔离保存。'),
            ('权限管理', '团队成员可按角色控制文件、知识库、模板和数据看板的查看与编辑权限。'),
            ('操作审计', '关键操作可记录处理人、时间、文件和任务状态，方便团队追溯。'),
            ('企业合规', '后续可接入企业安全策略、私有化部署、专属存储和合规审计需求。')
        ]
    )

@app.route('/resources/whitepaper')
def resource_whitepaper():
    return render_resource_page(
        '产品白皮书',
        '系统介绍序办的平台定位、核心能力、行业方案和企业部署思路。',
        [
            ('平台定位', '序办面向企业办公场景，连接智能表格、文档、云盘、知识库、自动化和数据分析。'),
            ('核心能力', '覆盖文件处理、表格清洗、自动报表、模板复用、团队协作和业务流程自动化。'),
            ('行业方案', '支持电商运营、财务、人事、销售、仓储、内容团队和企业数据分析场景。'),
            ('部署规划', '个人和团队可使用在线版，企业可按需规划 API 接入、权限体系和私有化部署。')
        ]
    )

@app.route('/knowledge')
def knowledge():
    return render_template('knowledge.html')

@app.route('/drive')
def drive():
    return render_template('drive.html')

# 价格路由
@app.route('/pricing')
def pricing():
    return render_template('pricing.html')

# 登录路由
@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/login/more')
def login_more():
    return render_template('login_more.html')

@app.route('/login/other')
def login_other():
    return render_template('login_other.html')

@app.route('/register')
def register():
    return render_template('register.html')

@app.route('/terms')
def terms():
    return render_template(
        'legal.html',
        title='用户协议',
        subtitle='使用序办前，请了解账号、文件处理、模板下载和平台服务规则。',
        sections=[
            ('账号与使用', '用户应妥善保管账号信息，不得使用平台上传、处理或传播违法违规内容。'),
            ('文件处理', '平台提供 AI 办公助手、文件上传、表格处理、清洗、转换和模板相关功能，处理结果请用户自行核对后使用。'),
            ('模板与下载', '模板中心内容用于办公参考，用户上传自定义模板时应确认拥有合法使用权。'),
            ('服务调整', '平台可能根据产品迭代调整功能、额度和会员权益，并在页面中展示最新说明。')
        ]
    )

@app.route('/privacy')
def privacy():
    return render_template(
        'legal.html',
        title='隐私政策',
        subtitle='我们重视用户数据和文件安全，本页说明平台如何处理账号信息和上传文件。',
        sections=[
            ('信息收集', '注册、登录、模板收藏和任务记录可能需要手机号、邮箱、昵称等基础账号信息。'),
            ('文件使用', '上传文件仅用于当前处理任务、预览和下载结果，不会用于与任务无关的用途。'),
            ('数据保存', '处理记录、下载记录和自定义模板会按账号保存，用户可在账号设置中管理相关信息。'),
            ('安全保护', '平台将持续完善访问控制、文件隔离和日志审计能力，降低数据泄露风险。')
        ]
    )

@app.route('/service')
def service_terms():
    return render_template(
        'legal.html',
        title='服务协议',
        subtitle='本协议用于说明平台功能范围、会员服务、服务支持和免责声明。',
        sections=[
            ('功能范围', '平台提供 AI 办公助手、智能表格、工具中心、模板中心、资源中心和企业协作能力展示。'),
            ('会员服务', '价格方案、处理次数、文件大小和团队权益以价格页面展示为准。'),
            ('服务支持', '用户可通过消息助手提交问题，工作人员会在工作时间内优先处理账号、文件和支付相关问题。'),
            ('免责声明', '平台会尽力保证处理稳定性，但复杂文件和数据处理结果仍需用户复核。')
        ]
    )

@app.route('/support')
def support():
    return redirect('/resources/help', code=302)

# 用户中心路由
@app.route('/workbench')
def workbench():
    return render_template('user_history.html')

@app.route('/console')
def console():
    return redirect('/workbench', code=302)

@app.route('/user/files')
def user_files():
    return render_template('user_files.html')

@app.route('/user/history')
def user_history():
    return render_template('user_history.html')

@app.route('/user/templates')
def user_templates():
    return render_template('user_templates.html')

def serialize_template_file(path: Path):
    stat = path.stat()
    name_parts = path.name.split('_', 1)
    original_name = name_parts[1] if len(name_parts) == 2 else path.name
    return {
        'filename': path.name,
        'name': Path(original_name).stem,
        'originalName': original_name,
        'size': stat.st_size,
        'savedAt': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
        'downloadUrl': f'/api/user/templates/{path.name}/download'
    }

@app.route('/api/user/templates', methods=['GET'])
def api_list_user_templates():
    template_dir = Path(app.config['TEMPLATE_FOLDER'])
    files = [
        serialize_template_file(path)
        for path in template_dir.iterdir()
        if path.is_file() and path.suffix.lower() in ALLOWED_TEMPLATE_EXTENSIONS
    ]
    files.sort(key=lambda item: item['savedAt'], reverse=True)
    return jsonify({'success': True, 'templates': files})

@app.route('/api/user/templates/upload', methods=['POST'])
def api_upload_user_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择模板文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '模板文件名为空'}), 400

    original_filename = file.filename
    ext = Path(original_filename).suffix.lower()
    if ext not in ALLOWED_TEMPLATE_EXTENSIONS:
        return jsonify({'success': False, 'error': f'不支持的模板格式: {ext}'}), 400

    safe_name = secure_filename(original_filename) or f'template{ext}'
    unique_name = f'{uuid.uuid4().hex}_{safe_name}'
    target_path = Path(app.config['TEMPLATE_FOLDER']) / unique_name
    file.save(target_path)

    logger.info(f"用户模板上传成功: {target_path}")
    return jsonify({
        'success': True,
        'template': serialize_template_file(target_path),
        'folder': str(Path(app.config['TEMPLATE_FOLDER']).resolve())
    })

@app.route('/api/user/templates/<filename>/download')
def api_download_user_template(filename):
    template_dir = Path(app.config['TEMPLATE_FOLDER']).resolve()
    target_path = (template_dir / filename).resolve()
    try:
        target_path.relative_to(template_dir)
    except ValueError:
        return jsonify({'success': False, 'error': '模板路径无效'}), 400

    if not target_path.is_file():
        return jsonify({'success': False, 'error': '模板不存在'}), 404

    return send_file(target_path, as_attachment=True)

@app.route('/api/user/templates/<filename>', methods=['DELETE'])
def api_delete_user_template(filename):
    template_dir = Path(app.config['TEMPLATE_FOLDER']).resolve()
    target_path = (template_dir / filename).resolve()
    try:
        target_path.relative_to(template_dir)
    except ValueError:
        return jsonify({'success': False, 'error': '模板路径无效'}), 400

    if not target_path.is_file():
        return jsonify({'success': False, 'error': '模板不存在'}), 404

    target_path.unlink()
    return jsonify({'success': True})

@app.route('/user/member')
def user_member():
    return render_template('user_member.html')

@app.route('/user/settings')
def user_settings():
    return render_template('user_settings.html')

@app.route('/api/columns', methods=['POST'])
def get_columns():
    """获取Excel文件的列名"""
    data = get_json_payload()
    filepath = data.get('filepath')
    auto_detect = bool(data.get('auto_detect', data.get('autoDetect', True)))
    try:
        header_row = coerce_positive_int(data.get('header_row', data.get('headerRow', 1)), '表头行')
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    
    logger.info(f"请求获取列名: filepath={filepath}, header_row={header_row}")
    
    if not filepath:
        logger.error("文件路径为空")
        return jsonify({'success': False, 'error': '文件路径为空'}), 400
    
    if not os.path.exists(filepath):
        logger.error(f"文件不存在: {filepath}")
        return jsonify({'success': False, 'error': f'文件不存在: {filepath}'}), 400

    if not is_uploaded_file(filepath):
        logger.error(f"非上传目录文件: {filepath}")
        return jsonify({'success': False, 'error': '只能读取已上传文件的列名'}), 400
    
    try:
        # 只读取表头
        wb = None
        if filepath.endswith('.xlsx') or filepath.endswith('.xlsm'):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        else:
            import xlrd
            wb = xlrd.open_workbook(filepath)
        
        try:
            if hasattr(wb, 'sheetnames'):
                sheet_names = wb.sheetnames
                ws = wb[sheet_names[0]]
            else:
                sheet_names = wb.sheet_names()
                ws = wb.sheet_by_index(0)
            
            logger.info(f"工作表名称: {sheet_names}")
            
            def read_header_values(row_number):
                if hasattr(ws, 'iter_rows'):
                    rows = list(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
                    return list(rows[0]) if rows else []

                row_index = row_number - 1
                if row_index < 0 or row_index >= ws.nrows:
                    return []
                return [ws.cell_value(row_index, col) for col in range(ws.ncols)]

            def clean_headers(values):
                return [str(value).strip() for value in values if value and str(value).strip()]

            effective_header_row = header_row
            non_empty_headers = clean_headers(read_header_values(header_row))

            if auto_detect and len(non_empty_headers) <= 1:
                best_header_row = header_row
                best_headers = non_empty_headers
                for candidate_row in range(header_row + 1, header_row + 11):
                    candidate_headers = clean_headers(read_header_values(candidate_row))
                    if len(candidate_headers) > len(best_headers):
                        best_header_row = candidate_row
                        best_headers = candidate_headers

                if len(best_headers) > len(non_empty_headers):
                    effective_header_row = best_header_row
                    non_empty_headers = best_headers
            
            logger.info(f"获取到列名: {non_empty_headers}")
            
            return jsonify({
                'success': True,
                'columns': non_empty_headers,
                'headerRow': effective_header_row
            })
        finally:
            if hasattr(wb, 'close'):
                wb.close()
            elif hasattr(wb, 'release_resources'):
                wb.release_resources()

    except Exception as e:
        logger.exception(f"获取列名失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    print("=" * 50)
    print("序办 Web服务")
    print("=" * 50)
    print("请访问: http://localhost:5000")
    print("按 Ctrl+C 停止服务")
    print("=" * 50)
    # 关闭调试器，避免 Werkzeug interactive console 占用业务路由。
    app.run(host='0.0.0.0', port=5000, debug=False, use_debugger=False, use_reloader=False)
